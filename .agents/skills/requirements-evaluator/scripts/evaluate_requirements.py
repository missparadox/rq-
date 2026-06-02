#!/usr/bin/env python3
from __future__ import annotations

import argparse
import importlib.util
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Sequence


DEFAULT_DIMENSIONS = [
    {
        "key": "or_user_language",
        "name": "OR-用户语言描述",
        "weight": 12,
        "description": "需求描述是否采用用户语言，避免技术术语，让非技术人员也能理解",
    },
    {
        "key": "or_scenario",
        "name": "OR-应用场景",
        "weight": 12,
        "description": "是否清晰描述了用户需求的应用场景，包括使用环境和上下文",
    },
    {
        "key": "or_user_value",
        "name": "OR-用户价值",
        "weight": 10,
        "description": "是否讲清楚解决的用户问题和带来的用户价值",
    },
    {
        "key": "or_constraints",
        "name": "OR-约束和限制",
        "weight": 6,
        "description": "是否明确约束和限制（如部署方式、合规性、性能指标等）",
    },
    {
        "key": "dr_security",
        "name": "DR-安全分析",
        "weight": 5,
        "description": "是否进行安全分析，描述对应的安全红线需求",
    },
    {
        "key": "dr_technical",
        "name": "DR-技术描述",
        "weight": 10,
        "description": "是否使用技术语言进行设计需求描述，功能描述清晰全面，包含正常情况和异常处理",
    },
    {
        "key": "dr_testability",
        "name": "DR-可测试性",
        "weight": 10,
        "description": "是否定量描述，对照 DR 能直接写出测试用例，并能推出关键验收条件与性能验证点",
    },
    {
        "key": "dr_ambiguity",
        "name": "DR-无歧义性",
        "weight": 8,
        "description": "参数规格是否清晰，无歧义",
    },
    {
        "key": "dr_exception",
        "name": "DR-异常描述",
        "weight": 7,
        "description": "是否明确描述异常路径、错误条件、无效输入处理、失败行为及边界场景",
    },
    {
        "key": "cross_scope",
        "name": "需求分解完整性",
        "weight": 7,
        "description": "是否覆盖 OR 的关键能力点，多个 DR 合起来是否形成完整分解，是否存在明显漏拆",
    },
    {
        "key": "cross_dependencies",
        "name": "需求分解边界清晰度",
        "weight": 6,
        "description": "各 DR 之间的职责边界是否清楚，是否存在交叉、重复拆解或责任不清",
    },
    {
        "key": "cross_traceability",
        "name": "需求映射一致性",
        "weight": 7,
        "description": "OR 与各 DR 是否语义一致、范围匹配、无明显偏题或冲突，且映射关系清楚稳定",
    },
]

CORE_FIELD_MAP = {
    "or_id": "OR需求编号",
    "or_name": "OR需求名称*",
    "or_desc": "OR需求描述*",
    "scenario": "应用场景",
    "customer_problem": "客户问题",
    "value_desc": "价值描述",
    "constraints": "约束与限制",
    "requirement_source": "需求来源",
    "region": "国家/区域",
    "or_category": "需求分类",
    "dr_id": "DR需求编号",
    "dr_name": "DR需求名称*",
    "dr_desc": "DR需求描述*",
    "dr_integration": "集成方式",
    "dr_param": "参数规格",
    "dr_operation": "操作场景",
    "dr_test": "系统测试要点",
    "dr_security": "安全约束",
    "spec_type": "规格分类",
    "subsystem": "所属子系统",
}

OR_CORE_ALIASES = (
    "or_id",
    "or_name",
    "or_desc",
    "scenario",
    "customer_problem",
    "value_desc",
    "constraints",
    "requirement_source",
    "region",
    "or_category",
)

DR_CORE_ALIASES = (
    "dr_id",
    "dr_name",
    "dr_desc",
    "dr_integration",
    "dr_param",
    "dr_operation",
    "dr_test",
    "dr_security",
    "spec_type",
    "subsystem",
)

DIMENSION_FIELD_MAP = {
    "or_user_language": ["OR需求名称*", "OR需求描述*", "更多描述信息"],
    "or_scenario": ["应用场景", "操作场景", "OR需求描述*", "更多描述信息"],
    "or_user_value": ["客户问题", "价值描述", "需求来源", "OR需求描述*"],
    "or_constraints": ["约束与限制", "国家/区域", "安全约束", "集成方式", "更多描述信息"],
    "dr_security": ["安全约束", "DR需求描述*"],
    "dr_technical": ["DR需求描述*", "集成方式", "参数规格", "规格分类", "所属子系统"],
    "dr_testability": ["系统测试要点", "参数规格", "DR需求描述*", "约束与限制", "更多描述信息"],
    "dr_ambiguity": ["参数规格", "DR需求描述*"],
    "dr_exception": ["DR需求描述*", "系统测试要点", "安全约束", "更多描述信息"],
    "cross_scope": ["OR需求描述*", "DR需求描述*"],
    "cross_dependencies": ["所属子系统", "国家/区域", "需求来源", "集成方式", "约束与限制"],
    "cross_traceability": ["OR需求编号", "DR需求编号", "ORURL", "DRURL"],
}

IGNORED_SOURCE_HEADERS = {
    "假设和依赖信息",
    "验证方法描述",
}

IGNORED_REQUIREMENT_PREFIXES = ("DS", "TS", "TDR", "TDS")
IGNORED_REQUIREMENT_SUFFIXES = ("需求编号", "需求名称", "需求描述", "URL")

SCORABLE_OR_CATEGORIES = {"功能"}

RED_LINE_RULES = [
    {
        "id": "R1",
        "name": "Missing User Value Cap",
        "rule": "OR 未显式写出用户价值、业务目的或合规动机",
        "cap": "OR部分 <= 24/40",
    },
    {
        "id": "R2",
        "name": "Untestable DR Cap",
        "rule": "DR 不能直接导出测试点、验收条件或验证步骤",
        "cap": "单个DR <= 24/40",
    },
    {
        "id": "R3",
        "name": "Vague DR Technical Description Cap",
        "rule": "DR 只有“支持/提供/实现”等模糊表述，缺少关键行为细节",
        "cap": "单个DR <= 20/40",
    },
    {
        "id": "R4",
        "name": "Incomplete Decomposition Coverage Cap",
        "rule": "OR 存在明显未覆盖的关键子能力",
        "cap": "分解与追踪质量 <= 10/20",
    },
    {
        "id": "R5",
        "name": "Distorted Decomposition Mapping Cap",
        "rule": "多个 DR 明显重复、冲突或无法追溯回 OR",
        "cap": "分解与追踪质量 <= 8/20",
    },
]

COMPACT_EVIDENCE_BOUNDARIES = [
    "OR四维仅使用OR核心字段和OR附加字段评分；DR细节不得补高OR，DR仅用于DR评分和OR-DR分解/映射。",
    "OR-用户语言描述：看OR是否以需求侧可理解的语言表达诉求；若主要是纯技术实现或方案细节，<=6/12。",
    "OR-约束和限制：看OR是否说明需求成立的范围、前提或限制；功能写得清楚但没有边界，<=2/6。",
]

COMPACT_DIMENSION_ANCHORS = {
    "or_user_language": "满足=用户需求清楚；缺失=无法识别用户需求",
    "or_scenario": "满足=角色/触发/环境明确；缺失=只有抽象能力",
    "or_user_value": "满足=价值或合规动机明确；缺失=仅功能描述",
    "or_constraints": "满足=关键范围或限制明确；缺失=无限制信息",
    "dr_security": "满足=安全控制明确；缺失=涉及安全但无控制",
    "dr_technical": "满足=行为/输入输出/条件清楚；缺失=仅支持/提供/实现",
    "dr_testability": "满足=可直接导出验收；缺失=无法验证",
    "dr_ambiguity": "满足=边界和结果明确；缺失=宽泛措辞",
    "dr_exception": "满足=失败处理明确；缺失=只有正常路径",
    "cross_scope": "满足=关键能力均覆盖；缺失=关键能力未落地",
    "cross_dependencies": "满足=DR边界清楚；缺失=边界混乱",
    "cross_traceability": "满足=OR/DR一致；缺失=映射冲突",
}

GRADING_BANDS = [
    {"grade": "优秀", "min_score": 90, "max_score": 100},
    {"grade": "良好", "min_score": 75, "max_score": 89},
    {"grade": "合格", "min_score": 60, "max_score": 74},
    {"grade": "待改进", "min_score": 0, "max_score": 59},
]


@dataclass
class RowRecord:
    index: int
    grouped: Dict[str, List[str]]

    def first(self, key: str, occurrence: int = 0) -> str:
        values = self.grouped.get(key, [])
        if occurrence < len(values):
            return clean_text(values[occurrence])
        return ""


@dataclass
class ReadResult:
    records: List[RowRecord]
    source_info: Dict[str, object]


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip())


def normalize_header_name(value: str) -> str:
    return clean_text(value).replace(" ", "")


def is_ignored_source_header(header: str) -> bool:
    normalized = normalize_header_name(header)
    if not normalized:
        return False
    if normalized in IGNORED_SOURCE_HEADERS:
        return True
    base = normalized.rstrip("*").upper()
    for prefix in IGNORED_REQUIREMENT_PREFIXES:
        for suffix in IGNORED_REQUIREMENT_SUFFIXES:
            if base == f"{prefix}{suffix}":
                return True
    return False


def filter_grouped_fields(grouped: Dict[str, List[str]]) -> Dict[str, List[str]]:
    return {
        key: values
        for key, values in grouped.items()
        if not is_ignored_source_header(key)
    }


def normalize_or_category(value: str) -> str:
    return clean_text(value) or "未分类"


def is_scorable_or_category(value: str) -> bool:
    return normalize_or_category(value) in SCORABLE_OR_CATEGORIES


def grade_from_score(score: float | int | None) -> str | None:
    if score is None:
        return None
    score_value = float(score)
    ordered_bands = sorted(GRADING_BANDS, key=lambda band: float(band["min_score"]))
    for index, band in enumerate(ordered_bands):
        min_score = float(band["min_score"])
        if score_value < min_score:
            continue
        next_band = ordered_bands[index + 1] if index + 1 < len(ordered_bands) else None
        if next_band is None:
            if score_value <= float(band["max_score"]):
                return str(band["grade"])
            return None
        if score_value < float(next_band["min_score"]):
            return str(band["grade"])
    return None


def build_compact_dimension_briefs(dimensions: Sequence[Dict[str, object]]) -> List[Dict[str, object]]:
    briefs = []
    for item in dimensions:
        key = str(item["key"])
        if key.startswith("or_"):
            applies_to = "OR"
        elif key.startswith("dr_"):
            applies_to = "DR"
        else:
            applies_to = "OR+全部关联DR"
        briefs.append(
            {
                "key": key,
                "name": item["name"],
                "weight": item["weight"],
                "applies_to": applies_to,
                "judgment_focus": item["description"],
                "anchor_summary": COMPACT_DIMENSION_ANCHORS[key],
            }
        )
    return briefs


def build_compact_rules(dimensions: Sequence[Dict[str, object]]) -> Dict[str, object]:
    score_weights = score_structure(dimensions)
    return {
        "scoring_formula": {
            "or_part": f"{score_weights['or_total_weight']}/100，每个OR只评一次",
            "dr_part": f"{score_weights['dr_total_weight']}/100，每个DR单独评分后取算术平均",
            "cross_part": f"{score_weights['cross_total_weight']}/100，按OR与全部关联DR的映射关系评分一次",
            "final_or_total": "OR部分 + DR平均分 + 需求分解与追踪质量",
        },
        "dimension_briefs": build_compact_dimension_briefs(dimensions),
        "evidence_rules": [
            "只根据当前OR、其关联DR以及包内保留的辅助字段中的显式证据打分，不得脑补。",
            "若文本仅引用外部需求、时序图或“如下图”，但当前包内未闭环，应按信息缺失处理。",
            "所有维度都必须返回，score 必须是数字；不允许输出 N/A。若证据不足或不适合作为独立要求呈现，直接给低分或 0 分并说明原因。",
        ],
        "evidence_boundaries": list(COMPACT_EVIDENCE_BOUNDARIES),
        "red_line_rules": [dict(rule) for rule in RED_LINE_RULES],
        "grading_bands": [dict(band) for band in GRADING_BANDS],
    }


def build_expected_output_schema(score_weights: Dict[str, int]) -> Dict[str, object]:
    return {
        "type": "object",
        "required": [
            "or_id",
            "or_name",
            "or_part",
            "dr_parts",
            "decomposition_quality",
            "review_conclusion",
            "review_decision",
            "triggered_red_line_rules",
            "key_evidence",
            "missing_items",
        ],
        "optional_but_supported": [],
        "notes": {
            "calculated_fields": "不输出汇总分或等级；聚合脚本将依据维度分重算 OR、DR平均、分解质量、总分和等级。",
            "lists": "blocking_issues、key_evidence、missing_items 各最多2条；无内容写“无”。",
            "dimension_scores": "所有维度必须完整返回，score 只能是数字；不允许使用 N/A/null 代替维度分数。",
            "dimension_reasons": "全量维度只返回分数；仅对得分率最低的2-3个维度单独说明。",
        },
    }


def detect_compact_warnings(raw_fields: Dict[str, List[str]]) -> List[str]:
    warnings = []
    flattened = " ".join(value for values in raw_fields.values() for value in values)
    if re.search(r"引用|如下图|时序图|见图|参考", flattened):
        warnings.append("当前OR包含外部引用或图示依赖；如果文本未闭环，不得因为外部引用而给高分。")
    if re.search(r"支持|提供|实现", flattened) and not re.search(r"默认|最大|最小|长度|范围|失败|异常|错误|校验", flattened):
        warnings.append("当前OR或DR可能存在大量口号式“支持/提供/实现”表述；需重点检查是否缺少行为细节。")
    if "系统测试要点" not in raw_fields:
        warnings.append("当前OR缺少直接测试字段；DR-可测试性需要从参数、条件和期望结果中谨慎判断。")
    return warnings


def build_dimensions() -> List[Dict[str, object]]:
    return [dict(item) for item in DEFAULT_DIMENSIONS]


def missing_runtime_dependencies(path: Path) -> List[str]:
    suffix = path.suffix.lower()
    missing = []
    if suffix in {".xlsx", ".xlsm"} and importlib.util.find_spec("openpyxl") is None:
        missing.append("openpyxl")
    return missing


def dependency_install_hint(packages: Sequence[str]) -> str:
    joined = " ".join(packages)
    executable = sys.executable or "python"
    if " " in executable and not (executable.startswith('"') and executable.endswith('"')):
        executable = f'"{executable}"'
    return f"{executable} -m pip install {joined}"


def ensure_runtime_dependencies(path: Path) -> None:
    missing = missing_runtime_dependencies(path)
    if not missing:
        return
    package_list = ", ".join(missing)
    hint = dependency_install_hint(missing)
    raise SystemExit(f"缺少运行依赖: {package_list}。请先执行: {hint}")


def read_excel(path: Path) -> ReadResult:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("读取 Excel 需要安装 openpyxl") from exc
    workbook = load_workbook(path, read_only=False, data_only=True)
    # 删除 sheet2（如果存在）
    if "Sheet2" in workbook.sheetnames:
        del workbook["Sheet2"]
    sheet = workbook.active
    source_info = {
        "input_format": path.suffix.lower().lstrip("."),
        "sheet_name": sheet.title,
    }
    if sheet.max_row < 1:
        return ReadResult(records=[], source_info=source_info)
    merged_values = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        anchor_value = sheet.cell(min_row, min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_values[(row, col)] = anchor_value

    def cell_value(row: int, col: int) -> str:
        value = sheet.cell(row, col).value
        if value is None and (row, col) in merged_values:
            value = merged_values[(row, col)]
        return "" if value is None else str(value)

    header = [cell_value(1, col) for col in range(1, sheet.max_column + 1)]
    while header and not clean_text(header[-1]):
        header.pop()
    if not header:
        return ReadResult(records=[], source_info=source_info)

    records = []
    for row_index in range(2, sheet.max_row + 1):
        grouped: Dict[str, List[str]] = defaultdict(list)
        values = [cell_value(row_index, col) for col in range(1, len(header) + 1)]
        for name, value in zip(header, values):
            grouped[name].append(value)
        records.append(RowRecord(index=row_index - 1, grouped=filter_grouped_fields(dict(grouped))))
    return ReadResult(records=records, source_info=source_info)


def read_json(path: Path) -> ReadResult:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        for value in data.values():
            if isinstance(value, list):
                data = value
                break
    if not isinstance(data, list):
        raise SystemExit("JSON 输入必须是对象数组，或包含数组字段的对象")
    records = []
    for idx, item in enumerate(data, start=1):
        if not isinstance(item, dict):
            continue
        grouped = filter_grouped_fields(
            {str(key): ["" if value is None else str(value)] for key, value in item.items()}
        )
        records.append(RowRecord(index=idx, grouped=grouped))
    return ReadResult(
        records=records,
        source_info={
            "input_format": path.suffix.lower().lstrip("."),
        },
    )


def read_records(path: Path) -> ReadResult:
    ensure_runtime_dependencies(path)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_excel(path)
    if suffix == ".json":
        return read_json(path)
    raise SystemExit(f"不支持的输入格式: {suffix}")


def summarize_headers(records: Sequence[RowRecord]) -> List[str]:
    if not records:
        return []
    headers = []
    seen = set()
    for record in records:
        for key in record.grouped:
            if key not in seen:
                seen.add(key)
                headers.append(key)
    return headers


def extract_core_fields(record: RowRecord) -> Dict[str, str]:
    core = {}
    for alias, source in CORE_FIELD_MAP.items():
        core[alias] = record.first(source)
    return core


def select_core_fields(core_fields: Dict[str, str], aliases: Sequence[str]) -> Dict[str, str]:
    return {alias: core_fields.get(alias, "") for alias in aliases}


def filter_dimensions(dimensions: Sequence[Dict[str, object]], prefix: str) -> List[Dict[str, object]]:
    return [dimension for dimension in dimensions if str(dimension["key"]).startswith(prefix)]


def score_structure(dimensions: Sequence[Dict[str, object]]) -> Dict[str, int]:
    return {
        "or_total_weight": sum(int(d["weight"]) for d in dimensions if str(d["key"]).startswith("or_")),
        "dr_total_weight": sum(int(d["weight"]) for d in dimensions if str(d["key"]).startswith("dr_")),
        "cross_total_weight": sum(int(d["weight"]) for d in dimensions if str(d["key"]).startswith("cross_")),
    }


def merge_records(records: Sequence[RowRecord]) -> RowRecord:
    grouped: Dict[str, List[str]] = defaultdict(list)
    for record in records:
        for key, values in record.grouped.items():
            for value in values:
                text = clean_text(value)
                if not text:
                    continue
                if text not in grouped[key]:
                    grouped[key].append(text)
    return RowRecord(index=records[0].index if records else 0, grouped=dict(grouped))


def build_raw_fields(record: RowRecord) -> Dict[str, List[str]]:
    return {
        key: [clean_text(value) for value in values if clean_text(value)]
        for key, values in record.grouped.items()
        if any(clean_text(value) for value in values)
    }


def build_or_review_skeleton(score_weights: Dict[str, int], dr_items: Sequence[Dict[str, object]]) -> Dict[str, object]:
    return {
        "or_total_score": {
            "max_score": score_weights["or_total_weight"] + score_weights["dr_total_weight"] + score_weights["cross_total_weight"],
            "score": None,
            "grade": None,
            "review_conclusion": None,
        },
        "or_part": {
            "max_score": score_weights["or_total_weight"],
            "score": None,
            "dimension_scores": [],
        },
        "dr_parts": [
            {
                "dr_id": item["id"],
                "dr_name": item["name"],
                "max_score": score_weights["dr_total_weight"],
                "score": None,
                "dimension_scores": [],
            }
            for item in dr_items
        ],
        "dr_average": {
            "max_score": score_weights["dr_total_weight"],
            "score": None,
        },
        "decomposition_quality": {
            "max_score": score_weights["cross_total_weight"],
            "score": None,
            "dimension_scores": [],
        },
        "review_decision": {
            "design_review_readiness": None,
            "development_readiness": None,
            "test_design_readiness": None,
            "blocking_issues": [],
            "triggered_red_line_rules": [],
        },
    }


def group_records_by_or(records: Sequence[RowRecord]) -> List[List[RowRecord]]:
    groups: List[List[RowRecord]] = []
    current: List[RowRecord] = []
    current_key = ""

    for record in records:
        core = extract_core_fields(record)
        or_key = core.get("or_id") or ""
        if current and or_key and or_key != current_key:
            groups.append(current)
            current = [record]
            current_key = or_key
            continue
        if current:
            current.append(record)
            if or_key:
                current_key = or_key
            continue
        current = [record]
        current_key = or_key or f"ROW-{record.index}"

    if current:
        groups.append(current)
    return groups


def group_records_by_dr(records: Sequence[RowRecord]) -> List[List[RowRecord]]:
    groups: List[List[RowRecord]] = []
    current: List[RowRecord] = []
    current_key = ""

    for record in records:
        core = extract_core_fields(record)
        dr_key = core.get("dr_id") or f"ROW-{record.index}"
        if current and dr_key != current_key:
            groups.append(current)
            current = [record]
            current_key = dr_key
            continue
        if current:
            current.append(record)
            continue
        current = [record]
        current_key = dr_key

    if current:
        groups.append(current)
    return groups


def build_dimension_view(record: RowRecord, dimensions: Sequence[Dict[str, object]]) -> Dict[str, Dict[str, object]]:
    view = {}
    for dimension in dimensions:
        key = str(dimension["key"])
        mapped_fields = DIMENSION_FIELD_MAP.get(key, [])
        evidence_fields = {}
        missing_fields = []
        for field in mapped_fields:
            values = [clean_text(value) for value in record.grouped.get(field, []) if clean_text(value)]
            if values:
                evidence_fields[field] = values
            else:
                missing_fields.append(field)
        view[key] = {
            "name": dimension["name"],
            "mapped_fields": mapped_fields,
            "evidence_fields": evidence_fields,
            "missing_fields": missing_fields,
        }
    return view


def build_review_packet(
    input_path: Path,
    dimensions: List[Dict[str, object]],
    records: Sequence[RowRecord],
    source_info: Dict[str, object] | None = None,
) -> Dict[str, object]:
    or_dimensions = filter_dimensions(dimensions, "or_")
    dr_dimensions = filter_dimensions(dimensions, "dr_")
    cross_dimensions = filter_dimensions(dimensions, "cross_")

    score_weights = score_structure(dimensions)
    groups = []
    total_dr_count = 0
    all_category_counts: Dict[str, int] = defaultdict(int)
    for or_records in group_records_by_or(records):
        merged_or_record = merge_records(or_records)
        raw_fields = build_raw_fields(merged_or_record)
        if not raw_fields:
            continue
        full_core_fields = extract_core_fields(merged_or_record)
        or_category = normalize_or_category(full_core_fields.get("or_category", ""))
        all_category_counts[or_category] += 1
        if not is_scorable_or_category(or_category):
            continue
        or_core_fields = select_core_fields(full_core_fields, OR_CORE_ALIASES)
        requirement_id = full_core_fields.get("or_id") or full_core_fields.get("dr_id") or f"ROW-{or_records[0].index}"
        requirement_name = full_core_fields.get("or_name") or full_core_fields.get("dr_name") or requirement_id

        dr_items = []
        for dr_records in group_records_by_dr(or_records):
            merged_dr_record = merge_records(dr_records)
            full_dr_core_fields = extract_core_fields(merged_dr_record)
            dr_core_fields = select_core_fields(full_dr_core_fields, DR_CORE_ALIASES)
            dr_id = full_dr_core_fields.get("dr_id") or f"ROW-{dr_records[0].index}"
            dr_name = full_dr_core_fields.get("dr_name") or dr_id
            dr_raw_fields = build_raw_fields(merged_dr_record)
            if not dr_raw_fields:
                continue
            dr_items.append(
                {
                    "row_indices": [record.index for record in dr_records],
                    "id": dr_id,
                    "name": dr_name,
                    "core_fields": dr_core_fields,
                    "dimension_view": build_dimension_view(merged_dr_record, dr_dimensions),
                    "raw_fields": dr_raw_fields,
                }
            )
        total_dr_count += len(dr_items)

        groups.append(
            {
                "row_indices": [record.index for record in or_records],
                "id": requirement_id,
                "name": requirement_name,
                "or_core_fields": or_core_fields,
                "or_dimension_view": build_dimension_view(merged_or_record, or_dimensions),
                "dr_items": dr_items,
                "dr_count": len(dr_items),
                "cross_dimension_view": build_dimension_view(merged_or_record, cross_dimensions),
                "review_skeleton": build_or_review_skeleton(score_weights, dr_items),
                "raw_fields": raw_fields,
            }
        )

    return {
        "input_path": str(input_path),
        "source_info": dict(source_info or {}),
        "score_structure": score_weights,
        "item_count": len(groups),
        "or_count": len(groups),
        "dr_count": total_dr_count,
        "dimension_count": len(dimensions),
        "dimensions": dimensions,
        "header_summary": summarize_headers(records),
        "groups": groups,
        "all_category_counts": dict(all_category_counts),
    }


def build_per_or_packets(review_packet: Dict[str, object]) -> Dict[str, object]:
    compact_rules = build_compact_rules(review_packet["dimensions"])
    expected_output_schema = build_expected_output_schema(review_packet["score_structure"])
    fallback_output_contract = {
        "format": "tagged_text",
        "markers": {
            "start": "[OR_RESULT_START]",
            "end": "[OR_RESULT_END]",
        },
        "required_scalar_fields": [
            "or_id",
            "or_name",
            "review_conclusion",
        ],
        "optional_scalar_fields": [],
        "required_list_fields": [
            "triggered_red_line_rules",
            "key_evidence",
            "missing_items",
            "blocking_issues",
        ],
        "dimension_line_patterns": [
            "or_dimension.<维度名>: <score>/<max> | <reason>",
            "dr_dimension.<dr_id>.<维度名>: <score>/<max> | <reason>",
            "cross_dimension.<维度名>: <score>/<max> | <reason>",
        ],
        "computed_by_aggregator": [
            "or_total_score",
            "or_part_score",
            "dr_score",
            "dr_average_score",
            "decomposition_quality_score",
            "grade",
        ],
        "dimension_reason_max_chars": 20,
        "list_item_limit": 2,
    }
    or_packets = []
    for group in review_packet["groups"]:
        or_packets.append(
            {
                "packet_type": "per_or_review_packet",
                "input_path": review_packet["input_path"],
                "source_info": review_packet.get("source_info", {}),
                "score_structure": review_packet["score_structure"],
                "compact_rules": compact_rules,
                "expected_output_schema": expected_output_schema,
                "fallback_output_contract": fallback_output_contract,
                "input_warnings": detect_compact_warnings(group.get("raw_fields", {})),
                "or_unit": {
                    "row_indices": group["row_indices"],
                    "id": group["id"],
                    "name": group["name"],
                    "or_core_fields": group["or_core_fields"],
                    "or_dimension_view": group["or_dimension_view"],
                    "dr_items": group["dr_items"],
                    "dr_count": group["dr_count"],
                    "cross_dimension_view": group["cross_dimension_view"],
                    "review_skeleton": group["review_skeleton"],
                    "raw_fields": group["raw_fields"],
                },
            }
        )
    return {
        "packet_type": "per_or_review_packets",
        "input_path": review_packet["input_path"],
        "source_info": review_packet.get("source_info", {}),
        "score_structure": review_packet["score_structure"],
        "dimension_count": review_packet["dimension_count"],
        "all_category_counts": review_packet.get("all_category_counts", {}),
        "or_count": review_packet["or_count"],
        "dr_count": review_packet["dr_count"],
        "compact_rules": compact_rules,
        "expected_output_schema": expected_output_schema,
        "fallback_output_contract": fallback_output_contract,
        "or_packets": or_packets,
    }


def render_per_or_packets_markdown(packet_doc: Dict[str, object]) -> str:
    lines = []
    lines.append("# 单OR评审任务包")
    lines.append("")
    lines.append("该文件适用于上下文受限模型。每次只取一个 OR 任务包进行评分，再汇总结果。")
    lines.append("")
    lines.append("## 数据概览")
    lines.append("")
    lines.append(f"- 输入文件: `{packet_doc['input_path']}`")
    for key, value in packet_doc.get("source_info", {}).items():
        lines.append(f"- {key}: `{value}`")
    lines.append(f"- OR条目数: {packet_doc['or_count']}")
    lines.append(f"- DR条目数: {packet_doc['dr_count']}")
    lines.append("")
    lines.append("## OR任务包列表")
    lines.append("")
    for index, or_packet in enumerate(packet_doc["or_packets"], start=1):
        unit = or_packet["or_unit"]
        lines.append(f"### OR {index}: {unit['id']} {unit['name']}")
        lines.append("")
        lines.append(f"- 覆盖行: {', '.join(str(idx) for idx in unit['row_indices'])}")
        lines.append(f"- DR数量: {unit['dr_count']}")
        if or_packet["input_warnings"]:
            lines.append("- 输入风险提示:")
            for warning in or_packet["input_warnings"]:
                lines.append(f"  - {warning}")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def _append_compact_field_block(lines: List[str], title: str, fields: Dict[str, str], *, prefix: str = "- ") -> None:
    non_empty = [(key, value) for key, value in fields.items() if clean_text(value)]
    if not non_empty:
        return
    lines.append(title)
    lines.append("")
    for key, value in non_empty:
        lines.append(f"{prefix}{key}: {value}")
    lines.append("")


def _build_compact_extra_fields(
    raw_fields: Dict[str, List[str]],
    *,
    excluded_headers: Sequence[str],
    max_fields: int = 8,
    max_value_chars: int = 160,
) -> List[tuple[str, str]]:
    excluded = set(excluded_headers)
    items: List[tuple[str, str]] = []
    for key, values in raw_fields.items():
        if key in excluded:
            continue
        cleaned_values = [clean_text(value) for value in values if clean_text(value)]
        if not cleaned_values:
            continue
        display = " | ".join(cleaned_values[:2])
        if len(display) > max_value_chars:
            display = display[: max_value_chars - 3].rstrip() + "..."
        items.append((key, display))
        if len(items) >= max_fields:
            break
    return items


def _dimension_names_by_scope(or_packet: Dict[str, object], scope: str) -> str:
    items = [item for item in or_packet["compact_rules"]["dimension_briefs"] if item["applies_to"] == scope]
    return " / ".join(f"{item['name']}({item['weight']})" for item in items)


def _append_compact_dimension_anchors(lines: List[str], or_packet: Dict[str, object]) -> None:
    lines.append("| 维度/上限 | 判定锚点 |")
    lines.append("| --- | --- |")
    for item in or_packet["compact_rules"]["dimension_briefs"]:
        lines.append(f"| {item['name']}/{item['weight']} | {item['anchor_summary']} |")


def render_single_or_packet_markdown(or_packet: Dict[str, object]) -> str:
    unit = or_packet["or_unit"]
    lines = []
    lines.append("# 单OR评审任务包")
    lines.append("")
    lines.append("本文件适用于上下文受限模型。当前输入只包含一个 OR 及其全部关联 DR。")
    lines.append("")
    lines.append("## 数据概览")
    lines.append("")
    lines.append(f"- 输入文件: `{or_packet['input_path']}`")
    for key, value in or_packet.get("source_info", {}).items():
        lines.append(f"- {key}: `{value}`")
    lines.append(f"- OR编号: `{unit['id']}`")
    lines.append(f"- OR名称: `{unit['name']}`")
    lines.append(f"- 覆盖行: {', '.join(str(idx) for idx in unit['row_indices'])}")
    lines.append(f"- DR数量: {unit['dr_count']}")
    if or_packet["input_warnings"]:
        lines.append("- 输入风险提示:")
        for warning in or_packet["input_warnings"]:
            lines.append(f"  - {warning}")
    lines.append("")
    lines.append("## 评分与输出约束")
    lines.append("")
    lines.append("- 总分: `OR(40) + DR平均(40) + 分解(20)`；逐维评分，汇总分和等级由脚本重算。")
    lines.append("- 证据: 只按本包显式内容；外部引用或缺失图示按信息缺失；所有维度必须返回数字分。")
    lines.append("- 输出: 仅输出下方 tagged 结果；全量维度只返回分数；仅对得分率最低的2-3个维度单独说明；结论一句话；各列表最多2条，无则写 `- 无`。")
    lines.append("")
    lines.append("### 证据边界")
    lines.append("")
    for boundary in or_packet["compact_rules"]["evidence_boundaries"]:
        lines.append(f"- {boundary}")
    lines.append("")
    _append_compact_dimension_anchors(lines, or_packet)
    lines.append("")
    lines.append("### 红线")
    lines.append("")
    for rule in or_packet["compact_rules"]["red_line_rules"]:
        lines.append(f"- {rule['id']}: {rule['rule']} -> `{rule['cap']}`")
    lines.append("")
    lines.append("## OR证据")
    lines.append("")
    _append_compact_field_block(lines, "### OR核心字段", unit["or_core_fields"])
    or_extra_fields = _build_compact_extra_fields(
        unit.get("raw_fields", {}),
        excluded_headers=[CORE_FIELD_MAP[alias] for alias in OR_CORE_ALIASES],
        max_fields=6,
    )
    if or_extra_fields:
        lines.append("### OR附加非空字段")
        lines.append("")
        for key, value in or_extra_fields:
            lines.append(f"- {key}: {value}")
        lines.append("")

    lines.append("## DR证据")
    lines.append("")
    for dr_item in unit["dr_items"]:
        lines.append(f"### DR {dr_item['id']} {dr_item['name']}")
        lines.append("")
        lines.append(f"- 覆盖行: {', '.join(str(idx) for idx in dr_item['row_indices'])}")
        dr_fields = {key: value for key, value in dr_item["core_fields"].items() if key.startswith("dr_")}
        for key, value in dr_fields.items():
            if value:
                lines.append(f"- {key}: {value}")
        dr_extra_fields = _build_compact_extra_fields(
            dr_item.get("raw_fields", {}),
            excluded_headers=[CORE_FIELD_MAP[alias] for alias in OR_CORE_ALIASES + DR_CORE_ALIASES],
            max_fields=8,
        )
        if dr_extra_fields:
            lines.append("- 附加非空字段:")
            for key, value in dr_extra_fields:
                lines.append(f"  - {key}: {value}")
        lines.append("")

    lines.append("## 输出骨架速记")
    lines.append("")
    lines.append("```text")
    lines.append("[OR_RESULT_START]")
    lines.append(f"or_id: {unit['id']}")
    lines.append(f"or_name: {unit['name']}")
    lines.append("or_dimension.OR-用户语言描述: <0-12>/<12>")
    lines.append("or_dimension.OR-应用场景: <0-12>/<12>")
    lines.append("or_dimension.OR-用户价值: <0-10>/<10>")
    lines.append("or_dimension.OR-约束和限制: <0-6>/<6>")
    for dr_item in unit["dr_items"]:
        lines.append(f"dr_dimension.{dr_item['id']}.DR-安全分析: <0-5>/<5>")
        lines.append(f"dr_dimension.{dr_item['id']}.DR-技术描述: <0-10>/<10>")
        lines.append(f"dr_dimension.{dr_item['id']}.DR-可测试性: <0-10>/<10>")
        lines.append(f"dr_dimension.{dr_item['id']}.DR-无歧义性: <0-8>/<8>")
        lines.append(f"dr_dimension.{dr_item['id']}.DR-异常描述: <0-7>/<7>")
    lines.append("cross_dimension.需求分解完整性: <0-7>/<7>")
    lines.append("cross_dimension.需求分解边界清晰度: <0-6>/<6>")
    lines.append("cross_dimension.需求映射一致性: <0-7>/<7>")
    lines.append("lowest_dimension_explanations:")
    lines.append("- OR.<维度名>: <score>/<max> | <扣分说明>")
    lines.append("- DR.<dr_id>.<维度名>: <score>/<max> | <扣分说明>")
    lines.append("- CROSS.<维度名>: <score>/<max> | <扣分说明>")
    lines.append("review_conclusion: <一句话结论>")
    lines.append("triggered_red_line_rules:")
    lines.append("- <规则编号或无>")
    lines.append("blocking_issues:")
    lines.append("- <阻塞项或无>")
    lines.append("key_evidence:")
    lines.append("- <证据1>")
    lines.append("missing_items:")
    lines.append("- <缺失项1>")
    lines.append("[OR_RESULT_END]")
    lines.append("```")
    lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def sanitize_filename_component(value: str) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9._-]+", "-", text)
    text = text.strip("-.")
    return text or "or"


def sanitize_artifact_dirname(value: str) -> str:
    text = clean_text(value).strip()
    text = re.sub(r'[\\/:*?"<>|]+', "-", text)
    text = text.strip(". ")
    return text or "requirement-file"


def get_project_root() -> Path:
    return Path(__file__).resolve().parents[4]


def build_artifact_dir(input_path: Path) -> Path:
    return get_project_root() / "reports" / sanitize_artifact_dirname(input_path.stem)


def resolve_default_output_path(
    input_path: Path,
    packet_scope: str,
    output_format: str,
    or_id: str | None = None,
) -> Path:
    artifact_dir = build_artifact_dir(input_path)
    extension = "json" if output_format == "json" else "md"
    if packet_scope == "all-or":
        return artifact_dir / "packets"
    if packet_scope == "single-or":
        target = sanitize_filename_component(or_id or input_path.stem)
        return artifact_dir / "packets" / f"single-{target}.{extension}"
    return artifact_dir / f"full-packet.{extension}"


def build_or_packet_filename(index: int, or_packet: Dict[str, object], extension: str) -> str:
    unit = or_packet["or_unit"]
    base = sanitize_filename_component(str(unit["id"]))
    return f"or-{index:03d}-{base}.{extension}"


def select_or_packet(packet_doc: Dict[str, object], or_id: str) -> Dict[str, object]:
    target = clean_text(or_id)
    for or_packet in packet_doc["or_packets"]:
        if clean_text(str(or_packet["or_unit"]["id"])) == target:
            return or_packet
    raise SystemExit(f"未找到 OR: {or_id}")


def write_packet_output(path: Path, packet: Dict[str, object], output_format: str) -> None:
    if output_format == "json":
        write_output(path, json.dumps(packet, ensure_ascii=False, indent=2))
        return
    write_output(path, render_single_or_packet_markdown(packet))


def write_split_packet_bundle(output_dir: Path, packet_doc: Dict[str, object], output_format: str) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    extension = "json" if output_format == "json" else "md"
    index_path = output_dir / f"index.{extension}"
    if output_format == "json":
        write_output(index_path, json.dumps(packet_doc, ensure_ascii=False, indent=2))
    else:
        write_output(index_path, render_per_or_packets_markdown(packet_doc))
    for index, or_packet in enumerate(packet_doc["or_packets"], start=1):
        filename = build_or_packet_filename(index, or_packet, extension)
        write_packet_output(output_dir / filename, or_packet, output_format)


def emit_split_packet_manifest(output_dir: Path, packet_doc: Dict[str, object], output_format: str) -> None:
    extension = "json" if output_format == "json" else "md"
    manifest = {
        "status": "ok",
        "packet_scope": "all-or",
        "packets_dir": str(output_dir),
        "index_path": str(output_dir / f"index.{extension}"),
        "packet_count": len(packet_doc["or_packets"]),
        "packet_files": [
            str(output_dir / build_or_packet_filename(index, or_packet, extension))
            for index, or_packet in enumerate(packet_doc["or_packets"], start=1)
        ],
    }
    # ASCII JSON is stable in terminals that cannot render non-ASCII filenames.
    print(json.dumps(manifest, ensure_ascii=True))


def render_review_packet_markdown(packet: Dict[str, object]) -> str:
    lines = []
    lines.append("# 需求评审任务包")
    lines.append("")
    lines.append("本文件不是评分结果，而是提供给大模型使用的评审输入材料。模型应根据 skill 与 rubric 自主评分并输出正式中文报告。")
    lines.append("")
    lines.append("## 数据概览")
    lines.append("")
    lines.append(f"- 输入文件: `{packet['input_path']}`")
    for key, value in packet.get("source_info", {}).items():
        lines.append(f"- {key}: `{value}`")
    lines.append(f"- OR条目数: {packet['or_count']}")
    lines.append(f"- DR条目数: {packet['dr_count']}")
    lines.append(f"- 维度数: {packet['dimension_count']}")
    all_cats = packet.get("all_category_counts", {})
    if all_cats:
        lines.append("- OR需求分类统计:")
        for cat, count in all_cats.items():
            tag = "" if is_scorable_or_category(cat) else " (已排除)"
            lines.append(f"  - {cat}: {count}个{tag}")
    lines.append("")
    lines.append("## 评分结构")
    lines.append("")
    lines.append(f"- OR部分: {packet['score_structure']['or_total_weight']}")
    lines.append(f"- DR部分: {packet['score_structure']['dr_total_weight']}，每个 OR 下的多个 DR 分别评分后取平均")
    lines.append(f"- 需求分解与追踪质量部分: {packet['score_structure']['cross_total_weight']}，每个 OR 只评一次")
    lines.append("")
    lines.append("## 表头摘要")
    lines.append("")
    for header in packet["header_summary"]:
        lines.append(f"- {header}")
    lines.append("")
    lines.append("## 评审维度")
    lines.append("")
    for item in packet["dimensions"]:
        desc = item.get("description", "")
        weight = item.get("weight", "")
        lines.append(f"- {item['name']} ({weight}): {desc}")
    lines.append("")
    lines.append("## OR评审单元")
    lines.append("")
    for item in packet["groups"]:
        lines.append(f"### OR {item['id']} {item['name']}")
        lines.append("")
        lines.append(f"- 覆盖行: {', '.join(str(idx) for idx in item['row_indices'])}")
        lines.append(f"- DR数量: {item['dr_count']}")
        lines.append("")
        lines.append("OR核心字段：")
        for key, value in item["or_core_fields"].items():
            if value:
                lines.append(f"- {key}: {value}")
        lines.append("")
        lines.append("OR维度视图：")
        for key, dimension in item.get("or_dimension_view", {}).items():
            lines.append(f"- {key} / {dimension['name']}")
            if dimension["evidence_fields"]:
                for field, values in dimension["evidence_fields"].items():
                    lines.append(f"  - evidence {field}: {' | '.join(values)}")
            if dimension["missing_fields"]:
                lines.append(f"  - missing: {', '.join(dimension['missing_fields'])}")
        lines.append("")
        lines.append("DR评审单元：")
        for dr_item in item["dr_items"]:
            lines.append(f"- DR {dr_item['id']} {dr_item['name']}")
            lines.append(f"  - row_indices: {', '.join(str(idx) for idx in dr_item['row_indices'])}")
            for key, value in dr_item["core_fields"].items():
                if value and key.startswith("dr_"):
                    lines.append(f"  - {key}: {value}")
            for key, dimension in dr_item.get("dimension_view", {}).items():
                lines.append(f"  - {key} / {dimension['name']}")
                if dimension["evidence_fields"]:
                    for field, values in dimension["evidence_fields"].items():
                        lines.append(f"    - evidence {field}: {' | '.join(values)}")
                if dimension["missing_fields"]:
                    lines.append(f"    - missing: {', '.join(dimension['missing_fields'])}")
        lines.append("")
        lines.append("需求分解与追踪维度视图：")
        for key, dimension in item.get("cross_dimension_view", {}).items():
            lines.append(f"- {key} / {dimension['name']}")
            if dimension["evidence_fields"]:
                for field, values in dimension["evidence_fields"].items():
                    lines.append(f"  - evidence {field}: {' | '.join(values)}")
            if dimension["missing_fields"]:
                lines.append(f"  - missing: {', '.join(dimension['missing_fields'])}")
        lines.append("")
        lines.append("评分骨架：")
        lines.append(f"- OR总分槽位: {item['review_skeleton']['or_total_score']['max_score']}")
        lines.append(f"- OR部分槽位: {item['review_skeleton']['or_part']['max_score']}")
        lines.append(f"- DR平均分槽位: {item['review_skeleton']['dr_average']['max_score']}")
        lines.append(f"- 需求分解与追踪质量槽位: {item['review_skeleton']['decomposition_quality']['max_score']}")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a requirement review packet from Excel or JSON input.")
    parser.add_argument("--input", required=True, help="Path to the requirement input file (.xlsx, .xlsm, .json)")
    parser.add_argument(
        "--output",
        help=(
            "Output path. Defaults to reports/<input-file-stem>/full-packet.<ext> "
            "or reports/<input-file-stem>/packets/ for --packet-scope=all-or"
        ),
    )
    parser.add_argument(
        "--format",
        choices=("json", "markdown"),
        default="markdown",
        help="Output packet format",
    )
    parser.add_argument(
        "--packet-scope",
        choices=("full", "single-or", "all-or"),
        default="full",
        help="Emit one full packet, one selected OR packet, or a split bundle of all OR packets",
    )
    parser.add_argument("--or-id", help="Target OR id when --packet-scope=single-or")
    return parser.parse_args(argv)


def write_output(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def main(argv: Sequence[str] | None = None) -> None:
    args = parse_args(argv)
    input_path = Path(args.input)
    output_path = (
        Path(args.output)
        if args.output
        else resolve_default_output_path(input_path, args.packet_scope, args.format, args.or_id)
    )

    result = read_records(input_path)
    dimensions = build_dimensions()
    packet = build_review_packet(
        input_path=input_path,
        dimensions=dimensions,
        records=result.records,
        source_info=result.source_info,
    )

    if args.packet_scope == "single-or":
        if not args.or_id:
            raise SystemExit("--or-id is required when --packet-scope=single-or")
        packet_doc = build_per_or_packets(packet)
        target_packet = select_or_packet(packet_doc, args.or_id)
        write_packet_output(output_path, target_packet, args.format)
        return

    if args.packet_scope == "all-or":
        packet_doc = build_per_or_packets(packet)
        if output_path.exists() and not output_path.is_dir():
            raise SystemExit("--output must be a directory path when --packet-scope=all-or")
        if output_path.suffix and not output_path.exists():
            raise SystemExit("--output must be a directory path when --packet-scope=all-or")
        write_split_packet_bundle(output_path, packet_doc, args.format)
        emit_split_packet_manifest(output_path, packet_doc, args.format)
        return

    if args.format == "json":
        write_output(output_path, json.dumps(packet, ensure_ascii=False, indent=2))
        return

    write_output(output_path, render_review_packet_markdown(packet))


if __name__ == "__main__":
    main()
