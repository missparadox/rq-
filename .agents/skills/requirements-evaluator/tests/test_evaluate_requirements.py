import importlib.util
import io
import json
import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path


MODULE_PATH = Path(__file__).resolve().parents[1] / "scripts" / "evaluate_requirements.py"


def load_module():
    spec = importlib.util.spec_from_file_location("requirements_eval_test", MODULE_PATH)
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


class RequirementsEvaluatorPacketTests(unittest.TestCase):
    def setUp(self):
        self.module = load_module()

    def test_build_dimensions_returns_default_rubric(self):
        dimensions = self.module.build_dimensions()
        by_key = {item["key"]: item for item in dimensions}
        self.assertEqual(by_key["or_user_language"]["weight"], 12)
        self.assertEqual(by_key["dr_technical"]["weight"], 10)
        self.assertEqual(by_key["cross_scope"]["weight"], 7)
        self.assertEqual(by_key["dr_ambiguity"]["weight"], 8)
        self.assertEqual(by_key["dr_exception"]["weight"], 7)
        self.assertEqual(sum(item["weight"] for item in dimensions if item["key"].startswith("or_")), 40)
        self.assertEqual(sum(item["weight"] for item in dimensions if item["key"].startswith("dr_")), 40)
        self.assertEqual(sum(item["weight"] for item in dimensions if item["key"].startswith("cross_")), 20)
        self.assertEqual(by_key["or_scenario"]["name"], "OR-应用场景")

    def test_build_review_packet_keeps_rows_and_core_fields(self):
        record_1 = self.module.RowRecord(
            index=1,
            grouped={
                "OR需求编号": ["DOR-1"],
                "OR需求名称*": ["网络检测与诊断"],
                "OR需求描述*": ["提供网络检测与维护功能。"],
                "需求分类": ["功能"],
                "需求来源": ["客户定制"],
                "国家/区域": ["中国"],
                "DR需求编号": ["DDR-1"],
                "DR需求名称*": ["Ping检测"],
                "DR需求描述*": ["支持 Ping 检测。"],
                "参数规格": ["次数 1-10"],
                "系统测试要点": ["校验 Ping 输入"],
                "DS需求名称*": ["Ping规格"],
                "规格分类": ["接口规格"],
                "所属子系统": ["网络管理"],
            },
        )
        record_2 = self.module.RowRecord(
            index=2,
            grouped={
                "OR需求编号": ["DOR-1"],
                "OR需求名称*": ["网络检测与诊断"],
                "OR需求描述*": ["提供网络检测与维护功能。"],
                "需求分类": ["功能"],
                "需求来源": ["客户定制"],
                "国家/区域": ["中国"],
                "DR需求编号": ["DDR-2"],
                "DR需求名称*": ["Telnet检测"],
                "DR需求描述*": ["支持 Telnet 检测。"],
                "参数规格": ["端口 1-65535"],
                "系统测试要点": ["校验 Telnet 输入"],
                "DS需求名称*": ["Telnet规格"],
                "规格分类": ["接口规格"],
                "所属子系统": ["网络管理"],
            },
        )

        packet = self.module.build_review_packet(
            input_path=Path("requirements.json"),
            dimensions=self.module.DEFAULT_DIMENSIONS,
            records=[record_1, record_2],
            source_info={"input_format": "json"},
        )

        self.assertEqual(packet["item_count"], 1)
        self.assertEqual(packet["or_count"], 1)
        self.assertEqual(packet["dr_count"], 2)
        self.assertEqual(packet["score_structure"]["or_total_weight"], 40)
        self.assertEqual(packet["score_structure"]["dr_total_weight"], 40)
        self.assertEqual(packet["score_structure"]["cross_total_weight"], 20)
        self.assertEqual(packet["source_info"]["input_format"], "json")
        self.assertEqual(packet["groups"][0]["id"], "DOR-1")
        self.assertIn("raw_fields", packet["groups"][0])
        self.assertIn("or_dimension_view", packet["groups"][0])
        self.assertIn("cross_dimension_view", packet["groups"][0])
        self.assertEqual(packet["groups"][0]["or_core_fields"]["requirement_source"], "客户定制")
        self.assertEqual(packet["groups"][0]["or_core_fields"]["region"], "中国")
        self.assertEqual(packet["groups"][0]["dr_count"], 2)
        self.assertEqual(packet["groups"][0]["review_skeleton"]["or_total_score"]["max_score"], 100)
        self.assertEqual(packet["groups"][0]["review_skeleton"]["or_part"]["max_score"], 40)
        self.assertEqual(packet["groups"][0]["review_skeleton"]["dr_average"]["max_score"], 40)
        self.assertEqual(packet["groups"][0]["review_skeleton"]["decomposition_quality"]["max_score"], 20)
        self.assertEqual(len(packet["groups"][0]["review_skeleton"]["dr_parts"]), 2)
        self.assertEqual(packet["groups"][0]["dr_items"][0]["core_fields"]["dr_desc"], "支持 Ping 检测。")
        self.assertEqual(packet["groups"][0]["dr_items"][0]["core_fields"]["spec_type"], "接口规格")
        self.assertEqual(packet["groups"][0]["dr_items"][0]["core_fields"]["subsystem"], "网络管理")
        self.assertIn("dr_testability", packet["groups"][0]["dr_items"][0]["dimension_view"])
        self.assertEqual(packet["groups"][0]["or_dimension_view"]["or_constraints"]["evidence_fields"]["国家/区域"], ["中国"])
        self.assertEqual(packet["groups"][0]["dr_items"][0]["dimension_view"]["dr_technical"]["evidence_fields"]["规格分类"], ["接口规格"])
        self.assertEqual(packet["groups"][0]["cross_dimension_view"]["cross_dependencies"]["evidence_fields"]["所属子系统"], ["网络管理"])

    def test_dimension_view_is_based_on_rubric_relevance_not_on_non_empty_frequency(self):
        record = self.module.RowRecord(
            index=1,
            grouped={
                "OR需求编号": ["DOR-2"],
                "OR需求名称*": ["设备接入"],
                "OR需求描述*": ["支持设备接入。"],
                "需求分类": ["功能"],
                "DR需求描述*": ["设备接入需校验参数格式。"],
                "参数规格": ["字段长度 1-64"],
            },
        )

        packet = self.module.build_review_packet(
            input_path=Path("sample.json"),
            dimensions=self.module.DEFAULT_DIMENSIONS,
            records=[record],
            source_info={"input_format": "json"},
        )
        dim_view = packet["groups"][0]["dr_items"][0]["dimension_view"]

        self.assertIn("dr_testability", dim_view)
        self.assertIn("系统测试要点", dim_view["dr_testability"]["mapped_fields"])
        self.assertIn("系统测试要点", dim_view["dr_testability"]["missing_fields"])
        self.assertEqual(dim_view["dr_testability"]["evidence_fields"]["参数规格"], ["字段长度 1-64"])

    def test_rendered_markdown_is_a_review_packet_not_a_scored_report(self):
        packet = {
            "input_path": "requirements.json",
            "source_info": {"input_format": "json", "sheet_name": "Sheet1"},
            "score_structure": {"or_total_weight": 40, "dr_total_weight": 40, "cross_total_weight": 20},
            "item_count": 1,
            "or_count": 1,
            "dr_count": 2,
            "dimension_count": 1,
            "dimensions": [{"key": "dr_technical", "name": "DR-技术描述", "weight": 10, "description": "desc"}],
            "header_summary": ["OR需求编号", "DR需求描述*"],
            "groups": [
                {
                    "row_indices": [1, 2],
                    "id": "DOR-1",
                    "name": "DNS配置",
                    "or_core_fields": {"or_desc": "支持 DNS"},
                    "or_dimension_view": {},
                    "dr_count": 2,
                    "dr_items": [
                        {
                            "row_indices": [1],
                            "id": "DDR-1",
                            "name": "Ping检测",
                            "core_fields": {"dr_desc": "IP 0-255"},
                            "dimension_view": {
                                "dr_technical": {
                                    "name": "DR-技术描述",
                                    "mapped_fields": ["DR需求描述*", "参数规格"],
                                    "evidence_fields": {"DR需求描述*": ["IP 0-255"]},
                                    "missing_fields": ["参数规格"],
                                }
                            },
                            "raw_fields": {"DR需求描述*": ["IP 0-255"]},
                        }
                    ],
                    "cross_dimension_view": {
                        "cross_dependencies": {
                            "name": "需求分解边界清晰度",
                            "mapped_fields": ["假设和依赖信息"],
                            "evidence_fields": {},
                            "missing_fields": ["假设和依赖信息"],
                        }
                    },
                    "review_skeleton": {
                        "or_total_score": {"max_score": 100, "score": None, "grade": None, "review_conclusion": None},
                        "or_part": {"max_score": 40, "score": None, "dimension_scores": []},
                        "dr_parts": [{"dr_id": "DDR-1", "dr_name": "Ping检测", "max_score": 40, "score": None, "dimension_scores": []}],
                        "dr_average": {"max_score": 40, "score": None},
                        "decomposition_quality": {"max_score": 20, "score": None, "dimension_scores": []},
                        "review_decision": {
                            "design_review_readiness": None,
                            "development_readiness": None,
                            "test_design_readiness": None,
                            "blocking_issues": [],
                            "triggered_red_line_rules": [],
                        },
                    },
                    "raw_fields": {"OR需求编号": ["DOR-1"], "DR需求描述*": ["IP 0-255"]},
                }
            ],
        }

        rendered = self.module.render_review_packet_markdown(packet)

        self.assertIn("# 需求评审任务包", rendered)
        self.assertIn("## 评分结构", rendered)
        self.assertIn("- OR部分: 40", rendered)
        self.assertIn("- DR数量: 2", rendered)
        self.assertIn("- sheet_name: `Sheet1`", rendered)
        self.assertIn("DR评审单元", rendered)
        self.assertIn("需求分解与追踪维度视图", rendered)
        self.assertIn("评分骨架", rendered)
        self.assertIn("- OR总分槽位: 100", rendered)
        self.assertNotIn("总分:", rendered)
        self.assertNotIn("等级:", rendered)

    def test_build_per_or_packets_includes_fallback_contract(self):
        record = self.module.RowRecord(
            index=1,
            grouped={
                "OR需求编号": ["DOR-1"],
                "OR需求名称*": ["网络检测与诊断"],
                "OR需求描述*": ["提供网络检测与维护功能。"],
                "需求分类": ["功能"],
                "DR需求编号": ["DDR-1"],
                "DR需求名称*": ["Ping检测"],
                "DR需求描述*": ["支持 Ping 检测。"],
            },
        )

        packet = self.module.build_review_packet(
            input_path=Path("requirements.json"),
            dimensions=self.module.DEFAULT_DIMENSIONS,
            records=[record],
            source_info={"input_format": "json"},
        )
        per_or = self.module.build_per_or_packets(packet)

        self.assertEqual(per_or["packet_type"], "per_or_review_packets")
        self.assertEqual(per_or["fallback_output_contract"]["format"], "tagged_text")
        self.assertEqual(per_or["or_packets"][0]["fallback_output_contract"]["markers"]["start"], "[OR_RESULT_START]")
        self.assertEqual(per_or["or_packets"][0]["expected_output_schema"]["optional_but_supported"], [])
        self.assertEqual(per_or["or_packets"][0]["fallback_output_contract"]["optional_scalar_fields"], [])
        self.assertNotIn("dr_score_line_pattern", per_or["or_packets"][0]["fallback_output_contract"])
        self.assertIn("dr_score", per_or["or_packets"][0]["fallback_output_contract"]["computed_by_aggregator"])

    def test_render_single_or_packet_markdown_contains_output_contract(self):
        record = self.module.RowRecord(
            index=1,
            grouped={
                "OR需求编号": ["DOR-1"],
                "OR需求名称*": ["网络检测与诊断"],
                "OR需求描述*": ["提供网络检测与维护功能。"],
                "需求分类": ["功能"],
                "DR需求编号": ["DDR-1"],
                "DR需求名称*": ["Ping检测"],
                "DR需求描述*": ["支持 Ping 检测。"],
            },
        )

        packet = self.module.build_review_packet(
            input_path=Path("requirements.json"),
            dimensions=self.module.DEFAULT_DIMENSIONS,
            records=[record],
            source_info={"input_format": "json"},
        )
        per_or = self.module.build_per_or_packets(packet)

        rendered = self.module.render_single_or_packet_markdown(per_or["or_packets"][0])

        self.assertIn("# 单OR评审任务包", rendered)
        self.assertIn("## 评分与输出约束", rendered)
        self.assertIn("| OR-用户价值/10 | 满足=价值或合规动机明确；缺失=仅功能描述 |", rendered)
        self.assertIn("| DR-可测试性/10 | 满足=可直接导出验收；缺失=无法验证 |", rendered)
        self.assertIn("| 需求映射一致性/7 | 满足=OR/DR一致；缺失=映射冲突 |", rendered)
        self.assertIn("### 红线", rendered)
        self.assertIn("R1: OR 未显式写出用户价值、业务目的或合规动机 -> `OR部分 <= 24/40`", rendered)
        self.assertIn("R5: 多个 DR 明显重复、冲突或无法追溯回 OR -> `分解与追踪质量 <= 8/20`", rendered)
        self.assertIn("仅输出下方 tagged 结果", rendered)
        self.assertIn("每条 `<短证据>` 建议不超过20字", rendered)
        self.assertIn("[OR_RESULT_START]", rendered)
        self.assertNotIn("dr_score.DDR-1", rendered)
        self.assertNotIn("OR总分槽位", rendered)
        self.assertIn("dr_dimension.DDR-1.DR-无歧义性: <0-8>/<8> | <短证据>", rendered)
        self.assertIn("## OR核心字段", rendered)

    def test_render_single_or_packet_markdown_keeps_dr_extra_non_empty_fields(self):
        record = self.module.RowRecord(
            index=1,
            grouped={
                "OR需求编号": ["DOR-1"],
                "OR需求名称*": ["网络检测与诊断"],
                "OR需求描述*": ["提供网络检测与维护功能。"],
                "需求分类": ["功能"],
                "DR需求编号": [""],
                "DR需求名称*": [""],
                "DR需求描述*": [""],
                "自定义DR说明": ["这里有真实的DR补充内容"],
                "验证补充": ["需要校验返回码"],
            },
        )

        packet = self.module.build_review_packet(
            input_path=Path("requirements.json"),
            dimensions=self.module.DEFAULT_DIMENSIONS,
            records=[record],
            source_info={"input_format": "json"},
        )
        per_or = self.module.build_per_or_packets(packet)

        rendered = self.module.render_single_or_packet_markdown(per_or["or_packets"][0])

        self.assertIn("附加非空字段", rendered)
        self.assertIn("自定义DR说明: 这里有真实的DR补充内容", rendered)
        self.assertIn("验证补充: 需要校验返回码", rendered)

    def test_build_review_packet_only_scores_functional_or_categories(self):
        functional = self.module.RowRecord(
            index=1,
            grouped={
                "OR需求编号": ["DOR-1"],
                "OR需求名称*": ["网络检测与诊断"],
                "OR需求描述*": ["提供网络检测与维护功能。"],
                "需求分类": ["功能"],
                "DR需求编号": ["DDR-1"],
                "DR需求名称*": ["Ping检测"],
                "DR需求描述*": ["支持 Ping 检测。"],
            },
        )
        non_functional = self.module.RowRecord(
            index=2,
            grouped={
                "OR需求编号": ["DOR-2"],
                "OR需求名称*": ["运维服务保障"],
                "OR需求描述*": ["提供运维保障。"],
                "需求分类": ["可服务性"],
                "DR需求编号": ["DDR-2"],
                "DR需求名称*": ["巡检要求"],
                "DR需求描述*": ["支持巡检。"],
            },
        )

        packet = self.module.build_review_packet(
            input_path=Path("requirements.json"),
            dimensions=self.module.DEFAULT_DIMENSIONS,
            records=[functional, non_functional],
            source_info={"input_format": "json"},
        )

        self.assertEqual(packet["or_count"], 1)
        self.assertEqual(packet["groups"][0]["id"], "DOR-1")
        self.assertEqual(packet["all_category_counts"]["功能"], 1)
        self.assertEqual(packet["all_category_counts"]["可服务性"], 1)

    def test_cli_json_output_writes_packet(self):
        rows = [
            {
                "OR需求编号": "DOR-1",
                "OR需求名称*": "网络检测与诊断",
                "OR需求描述*": "提供网络检测与维护功能。",
                "需求分类": "功能",
                "DR需求编号": "DDR-1",
                "DR需求名称*": "Ping检测",
                "DR需求描述*": "支持 Ping 检测。",
            },
            {
                "OR需求编号": "DOR-1",
                "OR需求名称*": "网络检测与诊断",
                "OR需求描述*": "提供网络检测与维护功能。",
                "需求分类": "功能",
                "DR需求编号": "DDR-2",
                "DR需求名称*": "Telnet检测",
                "DR需求描述*": "支持 Telnet 检测。",
            },
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            input_path = tmp / "requirements.json"
            output_path = tmp / "packet.json"
            input_path.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")

            self.module.main(
                [
                    "--input",
                    str(input_path),
                    "--output",
                    str(output_path),
                    "--format",
                    "json",
                ]
            )

            packet = json.loads(output_path.read_text(encoding="utf-8"))

        self.assertEqual(packet["item_count"], 1)
        self.assertEqual(packet["or_count"], 1)
        self.assertEqual(packet["dr_count"], 2)
        self.assertEqual(packet["source_info"]["input_format"], "json")
        self.assertEqual(packet["groups"][0]["name"], "网络检测与诊断")
        self.assertIn("review_skeleton", packet["groups"][0])

    def test_cli_single_or_json_output_writes_selected_packet(self):
        rows = [
            {
                "OR需求编号": "DOR-1",
                "OR需求名称*": "网络检测与诊断",
                "OR需求描述*": "提供网络检测与维护功能。",
                "需求分类": "功能",
                "DR需求编号": "DDR-1",
                "DR需求名称*": "Ping检测",
                "DR需求描述*": "支持 Ping 检测。",
            },
            {
                "OR需求编号": "DOR-2",
                "OR需求名称*": "端口检测",
                "OR需求描述*": "提供端口检测能力。",
                "需求分类": "功能",
                "DR需求编号": "DDR-2",
                "DR需求名称*": "Telnet检测",
                "DR需求描述*": "支持 Telnet 检测。",
            },
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            input_path = tmp / "requirements.json"
            output_path = tmp / "or.json"
            input_path.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")

            self.module.main(
                [
                    "--input",
                    str(input_path),
                    "--output",
                    str(output_path),
                    "--format",
                    "json",
                    "--packet-scope",
                    "single-or",
                    "--or-id",
                    "DOR-2",
                ]
            )

            packet = json.loads(output_path.read_text(encoding="utf-8"))

        self.assertEqual(packet["packet_type"], "per_or_review_packet")
        self.assertEqual(packet["or_unit"]["id"], "DOR-2")
        self.assertEqual(packet["or_unit"]["dr_count"], 1)

    def test_cli_all_or_markdown_output_writes_index_and_packets(self):
        rows = [
            {
                "OR需求编号": "DOR-1",
                "OR需求名称*": "网络检测与诊断",
                "OR需求描述*": "提供网络检测与维护功能。",
                "需求分类": "功能",
                "DR需求编号": "DDR-1",
                "DR需求名称*": "Ping检测",
                "DR需求描述*": "支持 Ping 检测。",
            },
            {
                "OR需求编号": "DOR-2",
                "OR需求名称*": "端口检测",
                "OR需求描述*": "提供端口检测能力。",
                "需求分类": "功能",
                "DR需求编号": "DDR-2",
                "DR需求名称*": "Telnet检测",
                "DR需求描述*": "支持 Telnet 检测。",
            },
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            input_path = tmp / "requirements.json"
            output_dir = tmp / "packets"
            input_path.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")

            stdout = io.StringIO()
            with redirect_stdout(stdout):
                self.module.main(
                    [
                        "--input",
                        str(input_path),
                        "--output",
                        str(output_dir),
                        "--packet-scope",
                        "all-or",
                    ]
                )

            index_text = (output_dir / "index.md").read_text(encoding="utf-8")
            packet_files = sorted(path.name for path in output_dir.glob("or-*.md"))
            manifest = json.loads(stdout.getvalue())

        self.assertIn("## OR任务包列表", index_text)
        self.assertEqual(packet_files, ["or-001-dor-1.md", "or-002-dor-2.md"])
        self.assertEqual(manifest["status"], "ok")
        self.assertEqual(manifest["packets_dir"], str(output_dir))
        self.assertEqual(manifest["index_path"], str(output_dir / "index.md"))
        self.assertEqual(manifest["packet_count"], 2)
        self.assertEqual(
            manifest["packet_files"],
            [str(output_dir / "or-001-dor-1.md"), str(output_dir / "or-002-dor-2.md")],
        )

    def test_cli_all_or_manifest_is_parseable_for_non_ascii_paths(self):
        rows = [
            {
                "OR需求编号": "DOR-1",
                "OR需求名称*": "网络检测与诊断",
                "OR需求描述*": "提供网络检测与维护功能。",
                "需求分类": "功能",
                "DR需求编号": "DDR-1",
                "DR需求名称*": "Ping检测",
                "DR需求描述*": "支持 Ping 检测。",
            }
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            input_path = tmp / "需求文档.json"
            output_dir = tmp / "报告" / "packets"
            input_path.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")

            stdout = io.StringIO()
            with redirect_stdout(stdout):
                self.module.main(
                    [
                        "--input",
                        str(input_path),
                        "--output",
                        str(output_dir),
                        "--packet-scope",
                        "all-or",
                    ]
                )

            raw_manifest = stdout.getvalue()
            manifest = json.loads(raw_manifest)

        self.assertNotIn("报告", raw_manifest)
        self.assertIn(r"\u62a5\u544a", raw_manifest)
        self.assertEqual(manifest["index_path"], str(output_dir / "index.md"))

    def test_cli_single_or_requires_or_id(self):
        rows = [
            {
                "OR需求编号": "DOR-1",
                "OR需求名称*": "网络检测与诊断",
                "OR需求描述*": "提供网络检测与维护功能。",
                "需求分类": "功能",
                "DR需求编号": "DDR-1",
                "DR需求名称*": "Ping检测",
                "DR需求描述*": "支持 Ping 检测。",
            }
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            input_path = tmp / "requirements.json"
            output_path = tmp / "or.json"
            input_path.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")

            with self.assertRaises(SystemExit) as ctx:
                self.module.main(
                    [
                        "--input",
                        str(input_path),
                        "--output",
                        str(output_path),
                        "--packet-scope",
                        "single-or",
                    ]
                )

        self.assertIn("--or-id is required", str(ctx.exception))

    def test_read_excel_records_active_sheet_name_in_source_info(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed in the current test environment")

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            input_path = tmp / "requirements.xlsx"

            workbook = openpyxl.Workbook()
            default_sheet = workbook.active
            default_sheet.title = "Sheet1"
            default_sheet["A1"] = "OR需求编号"
            default_sheet["B1"] = "OR需求名称*"
            default_sheet["C1"] = "OR需求描述*"
            default_sheet["D1"] = "DR需求编号"
            default_sheet["E1"] = "DR需求名称*"
            default_sheet["F1"] = "DR需求描述*"

            default_sheet["A2"] = "DOR-1"
            default_sheet["B2"] = "网络检测与诊断"
            default_sheet["C2"] = "提供网络检测与维护功能。"
            default_sheet["D2"] = "DDR-1"
            default_sheet["E2"] = "Ping检测"
            default_sheet["F2"] = "支持 Ping 检测。"
            default_sheet["D3"] = "DDR-2"
            default_sheet["E3"] = "Telnet检测"
            default_sheet["F3"] = "支持 Telnet 检测。"
            default_sheet.merge_cells("A2:A3")
            default_sheet.merge_cells("B2:B3")
            default_sheet.merge_cells("C2:C3")
            workbook.create_sheet("OtherSheet")
            workbook.save(input_path)

            result = self.module.read_excel(input_path)

        self.assertEqual(result.source_info["sheet_name"], "Sheet1")
        self.assertEqual(result.source_info["input_format"], "xlsx")
        self.assertEqual(len(result.records), 2)
        self.assertEqual(result.records[1].first("OR需求编号"), "DOR-1")
        self.assertEqual(result.records[1].first("OR需求名称*"), "网络检测与诊断")


    def test_missing_excel_dependency_message_contains_install_command(self):
        original_find_spec = self.module.importlib.util.find_spec

        def fake_find_spec(name):
            if name == "openpyxl":
                return None
            return original_find_spec(name)

        self.module.importlib.util.find_spec = fake_find_spec
        try:
            with self.assertRaises(SystemExit) as ctx:
                self.module.ensure_runtime_dependencies(Path("input.xlsx"))
        finally:
            self.module.importlib.util.find_spec = original_find_spec

        message = str(ctx.exception)
        self.assertIn("openpyxl", message)
        self.assertIn("python3 -m pip install openpyxl", message)


if __name__ == "__main__":
    unittest.main()
